#!/usr/bin/env python3
"""Auto-refresh dashboard from Google Sheets every hour on the hour."""

import subprocess, time, os, sys, requests, re, json
from datetime import datetime, timedelta

SHEET_ID = '1xqXetlPwD4f9pijzkJx7k5CBQqBJurFHtiP9IVX80K8'
DOWNLOAD_URL = f'https://docs.google.com/spreadsheets/d/{SHEET_ID}/export?format=xlsx'
OUTPUT_FILE = os.path.join(os.path.dirname(__file__), 'TNG - Báo cáo Vận Hành.xlsx')

HR_SHEET_ID = '1LSP9zDjsDzOOhkuSc3Tsz7df9qwl6xyJfFfGEygMYx0'
HR_DOWNLOAD_URL = f'https://docs.google.com/spreadsheets/d/{HR_SHEET_ID}/export?format=xlsx'
HR_OUTPUT_FILE = os.path.join(os.path.dirname(__file__), 'TNG_BÁO CÁO TUYỂN DỤNG.xlsx')

OPR_SHEET_ID = '1ip4a1c0nYtjdj8Y8TUM46pIYBp0tvoXY1XiMzG0a_Kc'
OPR_DOWNLOAD_URL = f'https://docs.google.com/spreadsheets/d/{OPR_SHEET_ID}/export?format=xlsx'
OPR_OUTPUT_FILE = os.path.join(os.path.dirname(__file__), 'TNG_OPR.xlsx')

BUILD_SCRIPT = os.path.join(os.path.dirname(__file__), 'build_dashboard.py')
EXTRACT_SCRIPT = os.path.join(os.path.dirname(__file__), 'extract_data.py')

WEATHER_CODES = {
    0: "Trời quang", 1: "Ít mây", 2: "Nhiều mây", 3: "Mây u ám",
    45: "Sương mù", 48: "Sương mù băng",
    51: "Mưa phùn nhẹ", 53: "Mưa phùn vừa", 55: "Mưa phùn dày",
    61: "Mưa nhẹ", 63: "Mưa vừa", 65: "Mưa to",
    80: "Mưa rào nhẹ", 81: "Mưa rào vừa", 82: "Mưa rào mạnh",
    95: "Giông bão nhẹ", 96: "Giông bão vừa", 99: "Giông bão mạnh"
}

def log(msg):
    print(f"[{datetime.now().strftime('%H:%M:%S')}] {msg}", flush=True)

def fetch_weather():
    """Fetch weather for 4 key provinces."""
    log("🌤️ Fetching weather data...")
    locations = {
        "Đắk Lắk": (12.6667, 108.05),
        "Gia Lai": (13.9833, 108.0),
        "Phú Yên": (13.0833, 109.3),
        "Bình Định": (13.7667, 109.2167)
    }
    weather_data = {}
    try:
        for name, coords in locations.items():
            url = f"https://api.open-meteo.com/v1/forecast?latitude={coords[0]}&longitude={coords[1]}&current_weather=true"
            r = requests.get(url, timeout=15)
            if r.status_code == 200:
                curr = r.json().get('current_weather', {})
                code = curr.get('weathercode')
                weather_data[name] = {
                    "temp": curr.get('temperature'),
                    "windspeed": curr.get('windspeed'),
                    "status": WEATHER_CODES.get(code, "Không xác định"),
                    "is_bad": code >= 61 if code is not None else False
                }
        
        with open(os.path.join(os.path.dirname(__file__), 'weather.json'), 'w', encoding='utf-8') as f:
            json.dump(weather_data, f, ensure_ascii=False, indent=2)
        log("✅ Weather data updated.")
    except Exception as e:
        log(f"❌ Weather fetch failed: {e}")

def start_chat_service():
    """Start Ngọc Trinh Chat Service if not running."""
    import socket
    s = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
    try:
        s.settimeout(1)
        s.connect(('127.0.0.1', 5005))
        s.close()
        log("🤖 Ngọc Trinh Chat Service is already running.")
    except:
        log("🚀 Starting Ngọc Trinh Chat Service...")
        subprocess.Popen([sys.executable, os.path.join(os.path.dirname(__file__), 'chat_service.py')], 
                         stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)

def download_sheet():
    """Download Google Sheets as xlsx."""
    log("📥 Downloading from Google Sheets...")
    success = True
    try:
        r = requests.get(DOWNLOAD_URL, timeout=30)
        if not r.content.startswith(b'PK'):
            raise ValueError("Downloaded content is not a valid Excel file (might be a login or permission error page).")
        with open(OUTPUT_FILE, 'wb') as f:
            f.write(r.content)
        log(f"✅ Downloaded Ops: {len(r.content):,} bytes")
    except Exception as e:
        log(f"❌ Download Ops failed: {e}")
        success = False
        
    try:
        r_hr = requests.get(HR_DOWNLOAD_URL, timeout=30)
        if not r_hr.content.startswith(b'PK'):
            raise ValueError("Downloaded content is not a valid Excel file (might be a login or permission error page).")
        with open(HR_OUTPUT_FILE, 'wb') as f:
            f.write(r_hr.content)
        log(f"✅ Downloaded HR: {len(r_hr.content):,} bytes")
    except Exception as e:
        log(f"❌ Download HR failed: {e}")
        success = False

    try:
        r_opr = requests.get(OPR_DOWNLOAD_URL, timeout=30)
        if not r_opr.content.startswith(b'PK'):
            raise ValueError("Downloaded content is not a valid Excel file (might be a login or permission error page).")
        with open(OPR_OUTPUT_FILE, 'wb') as f:
            f.write(r_opr.content)
        log(f"✅ Downloaded OPR: {len(r_opr.content):,} bytes")
    except Exception as e:
        log(f"❌ Download OPR failed: {e}")
        success = False
        
    return success

def run_extraction():
    """Re-extract all data from the downloaded xlsx."""
    import openpyxl, json
    
    def sn(v):
        try: return float(v) if v else 0
        except: return 0
    
    def is_closed_bc(v):
        if not v: return False
        s = str(v).strip()
        return '225 Phạm Văn Đồng' in s or '225 Pham Van Dong' in s

    def get_sheet_by_keyword(wb, keyword, exclude_keyword=None):
        for name in wb.sheetnames:
            if keyword.lower() in name.lower():
                if exclude_keyword and exclude_keyword.lower() in name.lower():
                    continue
                return wb[name]
        raise KeyError(f"Sheet with keyword '{keyword}' (excluding '{exclude_keyword}') not found.")
    
    wb = openpyxl.load_workbook(OUTPUT_FILE, data_only=True)
    
    # Extract actual report date from Excel (e.g. from '6.ONTIME TTS' cell I2)
    report_date = (datetime.now() - timedelta(days=1)).strftime('%Y-%m-%d')
    try:
        ws_date = get_sheet_by_keyword(wb, 'ONTIME TTS')
        date_str = str(ws_date.cell(2,9).value or '') # Cell I2
        match = re.search(r'(\d{4}-\d{2}-\d{2})', date_str)
        if match:
            report_date = match.group(1)
            log(f"📅 Report date from Excel: {report_date}")
        else:
            log(f"⚠️ No date match in Excel: {date_str}. Using N-1: {report_date}")
    except Exception as e:
        log(f"⚠️ Could not extract report date from Excel: {e}. Using N-1: {report_date}")

    data = {'report_date': report_date}
    
    try:
        ws = get_sheet_by_keyword(wb, 'GTC AM')
        gtc_tinh = []
        for r in range(5, 10):
            tinh = ws.cell(r,2).value
            if not tinh or 'Grand Total' in str(tinh): continue
            gtc_tinh.append({
                'tinh': str(tinh).strip(),
                'ca1_vol': sn(ws.cell(r,3).value), 'ca1_gan': sn(ws.cell(r,4).value), 'ca1_gtc': sn(ws.cell(r,5).value),
                'ca2_vol': sn(ws.cell(r,6).value), 'ca2_gan': sn(ws.cell(r,7).value), 'ca2_gtc': sn(ws.cell(r,8).value),
                'ton_vol': sn(ws.cell(r,9).value), 'ton_gan': sn(ws.cell(r,10).value), 'ton_gtc': sn(ws.cell(r,11).value),
                'total_vol': sn(ws.cell(r,12).value), 'total_gan': sn(ws.cell(r,13).value), 'total_gtc': sn(ws.cell(r,14).value),
            })
        data['gtc_tinh'] = gtc_tinh

        gtc_am = []
        for r in range(15, 60):
            am_val = ws.cell(r,2).value
            if not am_val: continue
            am_str = str(am_val).strip()
            if 'Grand Total' in am_str: break
            if 'Chi tiết' in am_str or 'Loại hàng' in am_str or 'BÁO CÁO' in am_str: continue
            
            gtc_am.append({
                'am': am_str,
                'ca1_vol': sn(ws.cell(r,3).value), 'ca1_gan': sn(ws.cell(r,4).value), 'ca1_gtc': sn(ws.cell(r,5).value),
                'ca2_vol': sn(ws.cell(r,6).value), 'ca2_gan': sn(ws.cell(r,7).value), 'ca2_gtc': sn(ws.cell(r,8).value),
                'ton_vol': sn(ws.cell(r,9).value), 'ton_gan': sn(ws.cell(r,10).value), 'ton_gtc': sn(ws.cell(r,11).value),
                'total_vol': sn(ws.cell(r,12).value), 'total_gan': sn(ws.cell(r,13).value), 'total_gtc': sn(ws.cell(r,14).value),
            })
        data['gtc_am'] = gtc_am
        
        gtc_tts_tinh = []
        for r in range(36, 41):
            tinh = ws.cell(r,2).value
            if not tinh: continue
            gtc_tts_tinh.append({
                'tinh': str(tinh).strip(),
                'ca1_vol': sn(ws.cell(r,3).value), 'ca1_gan': sn(ws.cell(r,4).value), 'ca1_gtc': sn(ws.cell(r,5).value),
                'ca2_vol': sn(ws.cell(r,6).value), 'ca2_gan': sn(ws.cell(r,7).value), 'ca2_gtc': sn(ws.cell(r,8).value),
                'ton_vol': sn(ws.cell(r,9).value), 'ton_gan': sn(ws.cell(r,10).value), 'ton_gtc': sn(ws.cell(r,11).value),
                'total_vol': sn(ws.cell(r,12).value), 'total_gan': sn(ws.cell(r,13).value), 'total_gtc': sn(ws.cell(r,14).value),
            })
        data['gtc_tts_tinh'] = gtc_tts_tinh

        gtc_tts_am = []
        for r in range(46, 61):
            am_val = ws.cell(r,2).value
            if not am_val: continue
            gtc_tts_am.append({
                'am': str(am_val).strip(),
                'ca1_vol': sn(ws.cell(r,3).value), 'ca1_gan': sn(ws.cell(r,4).value), 'ca1_gtc': sn(ws.cell(r,5).value),
                'ca2_vol': sn(ws.cell(r,6).value), 'ca2_gan': sn(ws.cell(r,7).value), 'ca2_gtc': sn(ws.cell(r,8).value),
                'ton_vol': sn(ws.cell(r,9).value), 'ton_gan': sn(ws.cell(r,10).value), 'ton_gtc': sn(ws.cell(r,11).value),
                'total_vol': sn(ws.cell(r,12).value), 'total_gan': sn(ws.cell(r,13).value), 'total_gtc': sn(ws.cell(r,14).value),
            })
        data['gtc_tts_am'] = gtc_tts_am
        data['grand_total_gtc'] = {'vol': sn(ws.cell(9,12).value), 'gan': sn(ws.cell(9,13).value), 'gtc': sn(ws.cell(9,14).value)}
        data['grand_total_gtc_tts'] = {'vol': sn(ws.cell(61,12).value), 'gan': sn(ws.cell(61,13).value), 'gtc': sn(ws.cell(61,14).value)}
    except Exception as e:
        log(f"⚠️ Error extracting GTC: {e}")

    try:
        ws = get_sheet_by_keyword(wb, 'LTC AM')
        ltc_am = []
        # Table 1: %LTC AM (Tổng ngày)
        for r in range(5, 22): # Data typically rows 5-20
            am = ws.cell(r,1).value
            if not am: continue
            am_str = str(am).strip()
            total_vol = sn(ws.cell(r,11).value)
            if total_vol == 0 and 'Grand' not in am_str: continue
            ltc_am.append({
                'am': am_str, 
                'ca1_vol': sn(ws.cell(r,2).value), 'ca1_gan': sn(ws.cell(r,3).value), 'ca1_ltc': sn(ws.cell(r,4).value),
                'ca2_vol': sn(ws.cell(r,5).value), 'ca2_gan': sn(ws.cell(r,6).value), 'ca2_ltc': sn(ws.cell(r,7).value),
                'ton_vol': sn(ws.cell(r,8).value), 'ton_gan': sn(ws.cell(r,9).value), 'ton_ltc': sn(ws.cell(r,10).value),
                'total_vol': total_vol, 'total_gan': sn(ws.cell(r,12).value), 'total_ltc': sn(ws.cell(r,13).value)
            })
            if 'Grand Total' in am_str: break
        data['ltc_am'] = ltc_am

        # Table 2: %LTC TTS AM (Lấy TTS)
        ltc_tts = []
        for r in range(26, 60): # Starts at row 26
            am = ws.cell(r,1).value
            if not am: continue
            am_str = str(am).strip()
            total_vol = sn(ws.cell(r,11).value)
            if total_vol == 0 and 'Grand' not in am_str: continue
            ltc_tts.append({
                'am': am_str, 
                'ca1_vol': sn(ws.cell(r,2).value), 'ca1_gan': sn(ws.cell(r,3).value), 'ca1_ltc': sn(ws.cell(r,4).value),
                'ca2_vol': sn(ws.cell(r,5).value), 'ca2_gan': sn(ws.cell(r,6).value), 'ca2_ltc': sn(ws.cell(r,7).value),
                'ton_vol': sn(ws.cell(r,8).value), 'ton_gan': sn(ws.cell(r,9).value), 'ton_ltc': sn(ws.cell(r,10).value),
                'total_vol': total_vol, 'total_gan': sn(ws.cell(r,12).value), 'total_ltc': sn(ws.cell(r,13).value)
            })
            if 'Grand Total' in am_str: break
        data['ltc_tts'] = ltc_tts
    except Exception as e: log(f"⚠️ Error extracting LTC AM/TTS: {e}")

    try:
        ws = get_sheet_by_keyword(wb, 'GTC-BC')
        gtc_bc = []
        for r in range(3, ws.max_row+1):
            am = ws.cell(r,2).value; bc = ws.cell(r,3).value
            if not am or not bc: continue
            if is_closed_bc(bc): continue
            total_vol = sn(ws.cell(r,13).value)
            if total_vol == 0: continue
            gtc_bc.append({'am': str(am).strip(), 'bc': str(bc).strip(), 'ca1_gtc': sn(ws.cell(r,6).value),
                           'ton_gtc': sn(ws.cell(r,12).value), 'total_vol': total_vol, 'total_gan': sn(ws.cell(r,14).value),
                           'total_gtc': sn(ws.cell(r,15).value), 'leadtime': sn(ws.cell(r,16).value) if ws.cell(r,16).value else 0})
        data['gtc_bc'] = gtc_bc
    except Exception as e: log(f"⚠️ Error extracting GTC BC: {e}")

    try:
        ws = get_sheet_by_keyword(wb, 'GTC TTS-BC')
        gtc_tts = []
        for r in range(3, ws.max_row+1):
            am = ws.cell(r,2).value; bc = ws.cell(r,3).value
            if not am or not bc: continue
            if is_closed_bc(bc): continue
            total_vol = sn(ws.cell(r,13).value)
            if total_vol == 0: continue
            gtc_tts.append({'am': str(am).strip(), 'bc': str(bc).strip(), 'ca1_gtc': sn(ws.cell(r,6).value),
                            'ton_gtc': sn(ws.cell(r,12).value), 'total_vol': total_vol, 'total_gan': sn(ws.cell(r,14).value),
                            'total_gtc': sn(ws.cell(r,15).value)})
        data['gtc_tts'] = gtc_tts
    except Exception as e: log(f"⚠️ Error extracting GTC TTS: {e}")

    try:
        ws = get_sheet_by_keyword(wb, 'ONTIME TTS')
        
        # Extract ontime dates from row 2
        ontime_dates = []
        for c in [3, 4, 5, 6, 7, 8, 9]:
            val = ws.cell(2, c).value
            if val:
                val_str = str(val).strip()
                match = re.search(r'(\d{4})-(\d{2})-(\d{2})', val_str)
                if match:
                    y, m, d = match.groups()
                    ontime_dates.append(f"{d}/{m}")
                else:
                    ontime_dates.append(val_str)
            else:
                ontime_dates.append("")
        data['ontime_dates'] = ontime_dates

        ontime_tts = []
        for r in range(3, ws.max_row+1):
            am = ws.cell(r,1).value
            if not am or 'AM' in str(am) or 'Bưu' in str(am): continue
            if is_closed_bc(am): continue
            try: today_val = float(ws.cell(r,9).value)
            except: continue
            ontime_tts.append({'am': str(am).strip(), 'day2': sn(ws.cell(r,2).value), 'day3': sn(ws.cell(r,3).value),
                               'day4': sn(ws.cell(r,4).value), 'day5': sn(ws.cell(r,5).value), 'day6': sn(ws.cell(r,6).value),
                               'day7': sn(ws.cell(r,7).value), 'day8': sn(ws.cell(r,8).value), 'today': today_val,
                               'n_change': sn(ws.cell(r,11).value)})
        data['ontime_tts'] = ontime_tts
    except Exception as e: log(f"⚠️ Error extracting Ontime TTS: {e}")

    try:
        ws = get_sheet_by_keyword(wb, 'BC Cảnh Báo', 'Vùng')
        canh_bao = []
        for r in range(4, ws.max_row+1):
            bc = ws.cell(r,3).value
            if not bc: continue
            if is_closed_bc(bc): continue
            canh_bao.append({
                'tinh': str(ws.cell(r,2).value or '').strip(), 'bc': str(bc).strip(), 
                'gtc_7d': sn(ws.cell(r,4).value), 'gtc_30d': sn(ws.cell(r,5).value), 
                'target': sn(ws.cell(r,6).value), 'n1': sn(ws.cell(r,7).value),
                'n2': sn(ws.cell(r,9).value), 'n3': sn(ws.cell(r,11).value)
            })
        data['canh_bao'] = canh_bao
    except Exception as e: log(f"⚠️ Error extracting Cảnh Báo: {e}")

    try:
        ws = get_sheet_by_keyword(wb, 'Cảnh Báo Vùng')
        canh_bao_vung = []
        for r in range(2, ws.max_row+1):
            bc = ws.cell(r,3).value
            if not bc: continue
            if is_closed_bc(bc): continue
            canh_bao_vung.append({
                'tinh': str(ws.cell(r,2).value or '').strip(), 'bc': str(bc).strip(), 
                'gtc_7d': sn(ws.cell(r,4).value), 'gtc_30d': sn(ws.cell(r,5).value), 
                'target': sn(ws.cell(r,7).value), 'gap': sn(ws.cell(r,8).value),
                'nhom': str(ws.cell(r,9).value or '')
            })
        data['canh_bao_vung'] = canh_bao_vung
    except Exception as e: log(f"⚠️ Error extracting Cảnh Báo Vùng: {e}")

    try:
        ws = get_sheet_by_keyword(wb, 'KINH DOANH')
        bc_kd_lay = []; total_lay = {}
        for r in range(4, ws.max_row+1):
            am = str(ws.cell(r,1).value or '').strip()
            if not am or am == 'AM': continue
            row_data = {
                'am': am, 'dt_days': [sn(ws.cell(r,c).value) for c in [2,4,6,8,10,12,14]],
                'vol_days': [sn(ws.cell(r,c).value) for c in [3,5,7,9,11,13,15]],
                'dt_total': sn(ws.cell(r,16).value), 'vol_total': sn(ws.cell(r,17).value),
                'n1_dt': sn(ws.cell(r,18).value), 'luyke': sn(ws.cell(r,20).value),
                'cungky': sn(ws.cell(r,21).value), 'so_cungky': sn(ws.cell(r,22).value),
            }
            if 'Tổng' in am: total_lay = row_data; break
            bc_kd_lay.append(row_data)
        data['bc_kd_lay'] = bc_kd_lay
        data['total_lay'] = total_lay
    except Exception as e: log(f"⚠️ Error extracting BC Kinh Doanh: {e}")

    # HR Extraction
    try:
        wb_ns = openpyxl.load_workbook(HR_OUTPUT_FILE, data_only=True)
        ws_nv = wb_ns['Báo cáo hàng ngày']
        am_groups = {}
        ns_bc = []
        ns_total = {'ptt_can': 0, 'ptt_co': 0, 'so_thieu': 0}
        for r in range(3, ws_nv.max_row+1):
            bc = str(ws_nv.cell(r,2).value or '').strip()
            if not bc or 'Grand Total' in bc: continue
            am = str(ws_nv.cell(r,3).value or '').strip()
            if not am: continue
            if is_closed_bc(bc): continue
            
            # Extract metrics
            db = sn(ws_nv.cell(r,11).value) # K: SL NVPTTT cần
            co = sn(ws_nv.cell(r,12).value) # L: SLNV đang có
            yctd = sn(ws_nv.cell(r,13).value) # M: YCTD
            da_tuyen = sn(ws_nv.cell(r,16).value) # P: Đã tuyển
            con_lam = sn(ws_nv.cell(r,17).value) # Q: Còn làm
            ct = sn(ws_nv.cell(r,19).value) # S: Cần tuyển
            dk_ob = sn(ws_nv.cell(r,18).value) # R: Dự kiến OB
            
            thieu = ct # Logic mới: Số thiếu = Cần tuyển (Cột S)
            
            row_bc = {
                'tinh': str(ws_nv.cell(r,1).value or '').strip(),
                'bc': bc, 'am': am, 'can': db, 'co': co, 'thieu': thieu,
                'yctd': yctd, 'da_tuyen': da_tuyen, 'con_lam': con_lam,
                'can_tuyen': ct, 'dk_ob': dk_ob
            }
            ns_bc.append(row_bc)
            
            if am not in am_groups:
                am_groups[am] = {
                    'am': am, 'ptt_can': 0, 'ptt_co': 0, 'so_thieu': 0,
                    'yctd': 0, 'da_tuyen': 0, 'con_lam': 0, 'dk_ob': 0, 'can_tuyen': 0
                }
            
            am_groups[am]['ptt_can'] += db
            am_groups[am]['ptt_co'] += co
            am_groups[am]['so_thieu'] += thieu
            am_groups[am]['yctd'] += yctd
            am_groups[am]['da_tuyen'] += da_tuyen
            am_groups[am]['con_lam'] += con_lam
            am_groups[am]['dk_ob'] += dk_ob
            am_groups[am]['can_tuyen'] += ct
            
            ns_total['ptt_can'] += db
            ns_total['ptt_co'] += co
            ns_total['so_thieu'] += thieu
        data['ns_am'] = sorted(am_groups.values(), key=lambda x: x['so_thieu'], reverse=True)
        data['ns_bc'] = ns_bc
        data['ns_total'] = ns_total
    except Exception as e: log(f"⚠️ HR Extraction failed: {e}")

    # OPR Extraction
    try:
        wb_opr = openpyxl.load_workbook(OPR_OUTPUT_FILE, data_only=True)
        ws_opr = wb_opr['DATA ']
        raw = []
        all_dates, all_procs, all_frames = set(), set(), set()
        
        def extract_tinh(chi_tiet):
            chi_tiet = str(chi_tiet or '').strip()
            if not chi_tiet:
                return None
            chi_tiet_norm = chi_tiet.replace('Đăk Lắk', 'Đắk Lắk').replace('Đăk Lăk', 'Đắk Lắk').replace('Đắc Lắc', 'Đắk Lắk')
            if '-' in chi_tiet_norm:
                tinh = chi_tiet_norm.split('-')[-1].strip()
                tinh = tinh.replace('Đăk Lắk', 'Đắk Lắk')
                return tinh
            for p in ['Bình Định', 'Đắk Lắk', 'Gia Lai', 'Phú Yên']:
                if p.lower() in chi_tiet_norm.lower():
                    return p
            return None

        for r in range(2, ws_opr.max_row + 1):
            prov = str(ws_opr.cell(r, 1).value or '').strip()
            chi_tiet = str(ws_opr.cell(r, 2).value or '').strip()
            frame = str(ws_opr.cell(r, 3).value or '').strip()
            date = str(ws_opr.cell(r, 4).value or '').strip()[:10]
            if not prov or prov == 'None': continue
            
            tinh = extract_tinh(chi_tiet) or 'Khác'
            item = {
                'prov': prov, 
                'tinh': tinh, 
                'frame': frame, 
                'date': date, 
                'vol_ltc': sn(ws_opr.cell(r, 5).value), 
                'vol': sn(ws_opr.cell(r, 7).value)
            }
            raw.append(item)
            all_dates.add(date)
            all_procs.add(prov)
            all_frames.add(frame)
        
        sorted_dates, sorted_procs, sorted_frames = sorted(list(all_dates)), sorted(list(all_procs)), sorted(list(all_frames))
        sorted_tinhs = ['Bình Định', 'Đắk Lắk', 'Gia Lai', 'Phú Yên']
        
        matrix_am = {p: {f: {d: {'vol_ltc':0, 'vol':0} for d in sorted_dates} for f in sorted_frames + ['Total']} for p in sorted_procs}
        matrix_tinh = {t: {f: {d: {'vol_ltc':0, 'vol':0} for d in sorted_dates} for f in sorted_frames + ['Total']} for t in sorted_tinhs}
        grand_total = {d: {'vol_ltc':0, 'vol':0} for d in sorted_dates}
        
        for item in raw:
            p, t, f, d = item['prov'], item['tinh'], item['frame'], item['date']
            # Update AM Matrix
            matrix_am[p][f][d]['vol_ltc'] += item['vol_ltc']; matrix_am[p][f][d]['vol'] += item['vol']
            matrix_am[p]['Total'][d]['vol_ltc'] += item['vol_ltc']; matrix_am[p]['Total'][d]['vol'] += item['vol']
            
            # Update Province Matrix
            if t in matrix_tinh:
                matrix_tinh[t][f][d]['vol_ltc'] += item['vol_ltc']; matrix_tinh[t][f][d]['vol'] += item['vol']
                matrix_tinh[t]['Total'][d]['vol_ltc'] += item['vol_ltc']; matrix_tinh[t]['Total'][d]['vol'] += item['vol']
                
            grand_total[d]['vol_ltc'] += item['vol_ltc']; grand_total[d]['vol'] += item['vol']
        
        opr_report = {'dates': sorted_dates, 'procs': []}
        for p in sorted_procs:
            proc_item = {'name': p, 'frames': []}
            for f in sorted_frames + ['Total']:
                fd = {'name': f, 'vals': {}}
                for d in sorted_dates:
                    d_data = matrix_am[p][f][d]
                    opr = d_data['vol']/d_data['vol_ltc'] if d_data['vol_ltc'] > 0 else 0
                    fd['vals'][d] = {'vol_ltc': d_data['vol_ltc'], 'opr': opr}
                proc_item['frames'].append(fd)
            opr_report['procs'].append(proc_item)
            
        opr_tinh_report = {'dates': sorted_dates, 'procs': []}
        for t in sorted_tinhs:
            proc_item = {'name': t, 'frames': []}
            for f in sorted_frames + ['Total']:
                fd = {'name': f, 'vals': {}}
                for d in sorted_dates:
                    d_data = matrix_tinh[t][f][d]
                    opr = d_data['vol']/d_data['vol_ltc'] if d_data['vol_ltc'] > 0 else 0
                    fd['vals'][d] = {'vol_ltc': d_data['vol_ltc'], 'opr': opr}
                proc_item['frames'].append(fd)
            opr_tinh_report['procs'].append(proc_item)
        
        target_date = sorted_dates[-1] if sorted_dates else datetime.now().strftime('%Y-%m-%d')
        opr_total = grand_total[target_date]['vol']/grand_total[target_date]['vol_ltc'] if target_date in grand_total and grand_total[target_date]['vol_ltc'] > 0 else 0
        data['opr_report'] = opr_report
        data['opr_tinh_report'] = opr_tinh_report
        data['opr_total'] = opr_total
        data['opr_daily'] = {d: (grand_total[d]['vol']/grand_total[d]['vol_ltc'] if grand_total[d]['vol_ltc'] > 0 else 0) for d in sorted_dates}
    except Exception as e: log(f"⚠️ OPR Extraction failed: {e}")

    # Save data
    with open(os.path.join(os.path.dirname(__file__), 'data.json'), 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2, default=str)
    with open(os.path.join(os.path.dirname(__file__), 'data.js'), 'w', encoding='utf-8') as f:
        f.write('window.globalData = ' + json.dumps(data, default=str) + ';')

    log(f"📊 Extraction completed: GTC={len(data.get('gtc_am',[]))}, ODR={len(data.get('ontime_tts',[]))}")
    return report_date

def update_history_json(report_date):
    """Update history.json and provincial history."""
    hist_path = os.path.join(os.path.dirname(__file__), 'history.json')
    prov_hist_file = os.path.join(os.path.dirname(__file__), 'gtc_prov_history.json')
    data_path = os.path.join(os.path.dirname(__file__), 'data.json')
    
    if not os.path.exists(data_path): return
    with open(data_path, 'r', encoding='utf-8') as f: data = json.load(f)
    
    ontime_vals = [float(x.get('today', 0)) for x in data.get('ontime_tts', []) if x.get('today') is not None]
    avg_ontime = sum(ontime_vals) / max(len(ontime_vals), 1)
    
    history = {}
    if os.path.exists(hist_path):
        with open(hist_path, 'r', encoding='utf-8') as f:
            try: history = json.load(f)
            except: history = {}
            
    # Dynamic correction: Update history.json with actual daily OPR values
    history_changed = False
    for d, opr_val in data.get('opr_daily', {}).items():
        if d in history:
            old_opr = history[d].get('opr')
            if old_opr is None or abs(old_opr - float(opr_val)) > 0.0001:
                history[d]['opr'] = float(opr_val)
                log(f"✍️ Corrected history OPR for {d}: {old_opr*100 if old_opr else 0:.2f}% -> {opr_val*100:.2f}%")
                history_changed = True
                
    if history_changed:
        with open(hist_path, 'w', encoding='utf-8') as f:
            json.dump(history, f, ensure_ascii=False, indent=2)
    
    # CUTOFF LOGIC: Comparison base (N-1) is frozen at the 23:00 snapshot.
    now = datetime.now()
    current_hour = now.hour
    is_new = report_date not in history
    is_cutoff = (current_hour == 23)
    
    if is_new or is_cutoff:
        gtc_v = float(data.get('grand_total_gtc', {}).get('gtc', 0))
        gtc_t = float(data.get('grand_total_gtc_tts', {}).get('gtc', 0))
        if gtc_t == 0 and gtc_v > 0:
            gtc_t = gtc_v + 0.005
            
        history[report_date] = {
            'vol': float(data.get('grand_total_gtc', {}).get('vol', 0)),
            'gtc_vung': gtc_v,
            'gtc_tts': gtc_t,
            'ontime': float(avg_ontime),
            'opr': float(data.get('opr_daily', {}).get(report_date, data.get('opr_total', 0))),
            'dt_luyke': float(data.get('total_lay', {}).get('luyke', 0)),
            'ns_thieu': float(data.get('ns_total', {}).get('so_thieu', 0)),
            'n_warn': len(data.get('canh_bao', []))
        }
        
        sorted_keys = sorted(history.keys())
        if len(sorted_keys) > 60:
            for k in sorted_keys[:-60]: del history[k]
            
        with open(hist_path, 'w', encoding='utf-8') as f:
            json.dump(history, f, ensure_ascii=False, indent=2)
        log(f"📈 History updated for {report_date} (Reason: {'New Date' if is_new else '23h Cutoff'})")
        # Update Provincial GTC History
        try:
            prov_hist = {}
            if os.path.exists(prov_hist_file):
                with open(prov_hist_file, 'r', encoding='utf-8') as f: prov_hist = json.load(f)
            prov_hist[report_date] = {x['tinh']: round(x['total_gtc']*100, 2) for x in data.get('gtc_tinh', [])}
            prov_hist[report_date]['Vùng TNG'] = round(data['grand_total_gtc']['gtc']*100, 2)
            with open(prov_hist_file, 'w', encoding='utf-8') as f: json.dump(prov_hist, f, indent=2, ensure_ascii=False)
            log(f"📈 Provincial History updated for {report_date}")
        except Exception as e: log(f"⚠️ Provincial history update failed: {e}")

        # Update Provincial GTC TTS History
        try:
            prov_tts_hist_file = os.path.join(os.path.dirname(__file__), 'gtc_tts_prov_history.json')
            prov_tts_hist = {}
            if os.path.exists(prov_tts_hist_file):
                with open(prov_tts_hist_file, 'r', encoding='utf-8') as f: prov_tts_hist = json.load(f)
            # Filter out 'Grand Total' or 'Vùng TNG' from gtc_tts_tinh and save
            prov_tts_hist[report_date] = {x['tinh']: round(x['total_gtc']*100, 2) for x in data.get('gtc_tts_tinh', []) if 'Grand' not in x['tinh'] and 'Vùng' not in x['tinh']}
            prov_tts_hist[report_date]['Vùng TNG'] = round(data['grand_total_gtc_tts']['gtc']*100, 2)
            with open(prov_tts_hist_file, 'w', encoding='utf-8') as f: json.dump(prov_tts_hist, f, indent=2, ensure_ascii=False)
            log(f"📈 Provincial GTC TTS History updated for {report_date}")
        except Exception as e: log(f"⚠️ Provincial GTC TTS history update failed: {e}")
    else:
        log(f"⏭️ History skip update for {report_date} (Current hour {current_hour} is not cutoff 23h)")

def extract_and_build():
    """Run build dashboard."""
    log("🏗️ Building dashboard...")
    result = subprocess.run([sys.executable, BUILD_SCRIPT], capture_output=True, text=True, cwd=os.path.dirname(__file__))
    if result.returncode == 0: log("✅ Dashboard built!")
    else: log(f"❌ Build failed: {result.stderr[:200]}")

def refresh():
    """Full refresh cycle."""
    fetch_weather()
    if download_sheet():
        try:
            if os.path.exists('data.json'):
                import shutil
                shutil.copy('data.json', 'data_snapshot.json')
            
            report_date = run_extraction()
            update_history_json(report_date)
            extract_and_build()
        except Exception as e:
            log(f"❌ Refresh failed: {e}")
            if os.path.exists('data_snapshot.json'):
                import shutil
                shutil.copy('data_snapshot.json', 'data.json')
                extract_and_build()
    else:
        log("⚠️ Using existing files due to download failure.")
        try:
            report_date = run_extraction()
            update_history_json(report_date)
        except Exception as e:
            log(f"❌ Extraction on existing files failed: {e}")
        extract_and_build()

if __name__ == '__main__':
    import argparse
    parser = argparse.ArgumentParser()
    parser.add_argument('--once', action='store_true', help='Run once and exit')
    args = parser.parse_args()

    is_github = os.environ.get('GITHUB_ACTIONS') == 'true'
    run_once = args.once or is_github

    log(f"🚀 Auto-refresh starting (Mode: {'Once' if run_once else 'Loop'})")
    
    # Skip chat service in GitHub Actions (no environment for it)
    if not is_github:
        start_chat_service()
    
    # Initial run
    refresh()
    
    # Sync to GitHub if running locally (GitHub Action workflow handles its own push)
    if not is_github:
        try:
            import sync_to_github
            sync_to_github.sync()
        except ImportError:
            log("⚠️ sync_to_github.py not found, skipping sync.")
        except Exception as e:
            log(f"⚠️ Sync failed: {e}")

    if run_once:
        log("✅ One-shot refresh completed.")
        sys.exit(0)
    
    last_run_hour = datetime.now().hour
    while True:
        now = datetime.now()
        if now.minute == 0 and now.hour != last_run_hour:
            log(f"⏰ Scheduled update: {now.hour}:00")
            refresh()
            
            # Sync after scheduled refresh
            if not is_github:
                try:
                    import sync_to_github
                    sync_to_github.sync()
                except: pass
                
            last_run_hour = now.hour
        time.sleep(30)
