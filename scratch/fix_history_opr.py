import openpyxl
import json
import os

def fix_opr_history():
    # 1. Parse OPR daily grand totals
    wb_opr = openpyxl.load_workbook('TNG_OPR.xlsx', data_only=True)
    ws_opr = wb_opr['DATA ']
    raw = []
    all_dates = set()
    for r in range(2, ws_opr.max_row + 1):
        prov = str(ws_opr.cell(r, 1).value or '').strip()
        date = str(ws_opr.cell(r, 4).value or '').strip()[:10]
        vol_ltc = ws_opr.cell(r, 5).value or 0
        vol = ws_opr.cell(r, 7).value or 0
        if not prov or prov == 'None': continue
        raw.append({'date': date, 'vol_ltc': vol_ltc, 'vol': vol})
        all_dates.add(date)

    sorted_dates = sorted(list(all_dates))
    grand_total = {d: {'vol_ltc':0, 'vol':0} for d in sorted_dates}
    for item in raw:
        grand_total[item['date']]['vol_ltc'] += item['vol_ltc']
        grand_total[item['date']]['vol'] += item['vol']

    opr_daily = {d: (grand_total[d]['vol']/grand_total[d]['vol_ltc'] if grand_total[d]['vol_ltc'] > 0 else 0) for d in sorted_dates}
    print("Computed actual daily OPR values:")
    for d, v in sorted(opr_daily.items()):
        print(f"  - {d}: {v*100:.2f}%")

    # 2. Load existing history
    history = {}
    if os.path.exists('history.json'):
        with open('history.json', 'r', encoding='utf-8') as f:
            history = json.load(f)

    # 3. Update OPR in history
    updated_count = 0
    for d, val in opr_daily.items():
        if d in history:
            old_val = history[d].get('opr')
            history[d]['opr'] = float(val)
            print(f"Updated history for {d}: OPR {old_val*100 if old_val else 0:.2f}% -> {val*100:.2f}%")
            updated_count += 1

    # 4. Save history
    with open('history.json', 'w', encoding='utf-8') as f:
        json.dump(history, f, indent=2, ensure_ascii=False)
    print(f"✅ Successfully corrected {updated_count} history entries.")

if __name__ == '__main__':
    fix_opr_history()
