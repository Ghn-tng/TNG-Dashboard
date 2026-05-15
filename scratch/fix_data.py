import openpyxl, json, os

def sn(v):
    try: return float(v) if v else 0
    except: return 0

def fix_data():
    wb = openpyxl.load_workbook('TNG - Báo cáo Vận Hành.xlsx', data_only=True)
    ws = wb['1. % GTC AM']
    
    # Extract ALL AMs from rows 15-29
    gtc_am = []
    for r in range(15, 30):
        am = ws.cell(r,2).value
        if not am: continue
        gtc_am.append({
            'am': str(am).strip(),
            'ca1_vol': sn(ws.cell(r,3).value), 'ca1_gan': sn(ws.cell(r,4).value), 'ca1_gtc': sn(ws.cell(r,5).value),
            'ca2_vol': sn(ws.cell(r,6).value), 'ca2_gan': sn(ws.cell(r,7).value), 'ca2_gtc': sn(ws.cell(r,8).value),
            'ton_vol': sn(ws.cell(r,9).value), 'ton_gan': sn(ws.cell(r,10).value), 'ton_gtc': sn(ws.cell(r,11).value),
            'total_vol': sn(ws.cell(r,12).value), 'total_gan': sn(ws.cell(r,13).value), 'total_gtc': sn(ws.cell(r,14).value),
        })
    
    # Extract Province Totals for Tab 1 table
    gtc_tinh = []
    for r in range(5, 10):
        tinh = ws.cell(r,2).value
        if not tinh: continue
        gtc_tinh.append({
            'tinh': str(tinh).strip(),
            'ca1_vol': sn(ws.cell(r,3).value), 'ca1_gan': sn(ws.cell(r,4).value), 'ca1_gtc': sn(ws.cell(r,5).value),
            'ca2_vol': sn(ws.cell(r,6).value), 'ca2_gan': sn(ws.cell(r,7).value), 'ca2_gtc': sn(ws.cell(r,8).value),
            'ton_vol': sn(ws.cell(r,9).value), 'ton_gan': sn(ws.cell(r,10).value), 'ton_gtc': sn(ws.cell(r,11).value),
            'total_vol': sn(ws.cell(r,12).value), 'total_gan': sn(ws.cell(r,13).value), 'total_gtc': sn(ws.cell(r,14).value),
        })

    with open('data.json', 'r', encoding='utf-8') as f:
        data = json.load(f)
    
    data['gtc_am'] = gtc_am
    data['gtc_tinh'] = gtc_tinh
    
    with open('data.json', 'w', encoding='utf-8') as f:
        json.dump(data, f, indent=2, ensure_ascii=False)
    print("Fixed data.json with all 15 AMs.")

if __name__ == "__main__":
    fix_data()
