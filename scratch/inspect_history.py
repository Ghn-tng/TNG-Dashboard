import openpyxl
import json
import os

def sn(v):
    if v is None: return 0
    try: return float(v)
    except: return 0

def inspect_all_history():
    wb_ops = openpyxl.load_workbook('TNG - Báo cáo Vận Hành.xlsx', data_only=True)
    ws_kd = wb_ops['11. BC KINH DOANH']
    
    # Find the row with 'Grand Total' or 'Tổng'
    for r in range(1, 40):
        val = str(ws_kd.cell(r, 1).value)
        if 'Grand Total' in val or 'Tổng' in val:
            row_vals = [ws_kd.cell(r, c).value for c in range(1, 25)]
            print(f"Total Row {r}: {row_vals}")

    wb_ops.close()

if __name__ == "__main__":
    inspect_all_history()
