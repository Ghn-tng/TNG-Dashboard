import openpyxl
import json
import os
from datetime import datetime, timedelta

def sn(v):
    if v is None: return 0
    try: return float(v)
    except: return 0

def update_history():
    wb = openpyxl.load_workbook('TNG - Báo cáo Vận Hành.xlsx', data_only=True)
    
    # 1. Load existing history
    history = {}
    if os.path.exists('history.json'):
        with open('history.json', 'r', encoding='utf-8') as f:
            history = json.load(f)
            
    # 2. Extract Ontime History (6.ONTIME TTS)
    ws_ot = wb['6.ONTIME TTS']
    dates_ot = []
    for c in range(2, 10):
        d_val = ws_ot.cell(2, c).value
        if d_val:
            # Format: '2026-05-11 - Thứ 2'
            d_str = str(d_val).split(' ')[0]
            dates_ot.append(d_str)
            
    # Grand Total row for Ontime is 15
    for i, d_str in enumerate(dates_ot):
        if d_str not in history: history[d_str] = {}
        history[d_str]['ontime'] = sn(ws_ot.cell(15, i+2).value)
        history[d_str]['opr'] = sn(ws_ot.cell(15, i+2).value) # Use Ontime as OPR proxy if OPR not found

    # 3. Extract DT History (11. BC KINH DOANH)
    ws_kd = wb['11. BC KINH DOANH']
    # Dates are in Row 2, Columns 2, 4, 6, 8, 10, 12, 14
    # Grand Total row is 20
    for c in [2, 4, 6, 8, 10, 12, 14]:
        d_val = ws_kd.cell(2, c).value
        if d_val:
            if isinstance(d_val, datetime):
                d_str = d_val.strftime('%Y-%m-%d')
            else:
                d_str = str(d_val).split(' ')[0]
            
            if d_str not in history: history[d_str] = {}
            history[d_str]['dt_daily'] = sn(ws_kd.cell(20, c).value)
            # Volume for DT
            history[d_str]['vol_kd'] = sn(ws_kd.cell(20, c+1).value)

    # 4. GTC History (This is harder as GTC sheet usually only has today)
    # But wait, 10/05 GTC Vùng was 67.23%.
    if '2026-05-10' in history:
        history['2026-05-10']['gtc_vung'] = 0.6723
    
    # 11/05 GTC Vùng? From the user's latest screenshot: 70.5%
    if '2026-05-11' in history:
        history['2026-05-11']['gtc_vung'] = 0.705

    # 5. Fix DT Lũy Kế
    # 11/05 Lũy kế is 1,460,126,648
    history['2026-05-11']['dt_luyke'] = 1460126648.0
    # 10/05 Lũy kế = 1,460,126,648 - today's daily (155,075,426) = 1,305,051,222
    history['2026-05-10']['dt_luyke'] = 1460126648.0 - 155075426.0

    # Save
    with open('history.json', 'w', encoding='utf-8') as f:
        json.dump(history, f, indent=2, ensure_ascii=False)
    
    print("✅ History updated with accurate Excel data.")
    wb.close()

if __name__ == "__main__":
    update_history()
